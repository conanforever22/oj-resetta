import json
import os
import re
import sys
from datetime import datetime, timezone
from openpyxl import load_workbook

# =====================================================================
# OJ 配置 — 新增 OJ 时在此添加一条
# =====================================================================

SHEET_OJ_MAP = {
    "洛谷":       "luogu",
    "Codeforces": "codeforces",
    "AtCoder":    "atcoder",
}

REMOTE_JUDGE_PREFIXES = {
    "AT_": "atcoder",
    "CF":  "codeforces",
}

EXCEL_FILE = "mappings.xlsx"
OUTPUT_JSON = "data.json"
OUTPUT_CSV_DIR = "csv"


# =====================================================================
# AtCoder 源 URL 生成
#   "abc123d"  → https://atcoder.jp/contests/abc123/tasks/abc123_d
#   "abc123d1" → https://atcoder.jp/contests/abc123/tasks/abc123_d1
#   "abc0101"  → https://atcoder.jp/contests/abc010/tasks/abc010_1
# =====================================================================
def _atcoder_source_url(source_id):
    normalized = source_id.lower().replace("_", "")
    m = re.match(r"^([a-z]+\d{1,3})([a-z].*|[1-8])$", normalized)
    if m:
        contest = m.group(1)
        task_suffix = m.group(2)
        return f"https://atcoder.jp/contests/{contest}/tasks/{contest}_{task_suffix}"
    return None


# =====================================================================
# Codeforces 源 URL 生成
#   "cf2038a"  → https://codeforces.com/problemset/problem/2038/A
#   "cf2038c1" → https://codeforces.com/problemset/problem/2038/C1
# =====================================================================
def _cf_source_url(source_id):
    m = re.match(r"^cf(\d+)([a-z]\d*)$", source_id.lower())
    if m:
        return f"https://codeforces.com/problemset/problem/{m.group(1)}/{m.group(2).upper()}"
    return None


# =====================================================================
# 源 URL 生成函数注册表
# =====================================================================
SOURCE_URL_BUILDERS = {
    "luogu":       lambda sid: f"https://www.luogu.com.cn/problem/{sid}",
    "codeforces":  _cf_source_url,
    "atcoder":     _atcoder_source_url,
}


# =====================================================================
# 构建主流程
# =====================================================================
def build():
    if not os.path.exists(EXCEL_FILE):
        print(f"Error: {EXCEL_FILE} not found")
        sys.exit(1)

    os.makedirs(OUTPUT_CSV_DIR, exist_ok=True)

    wb = load_workbook(EXCEL_FILE, data_only=True)

    by_source = {}
    by_my_id = {}
    total = 0

    for sheet_name, oj_key in SHEET_OJ_MAP.items():
        if sheet_name not in wb.sheetnames:
            print(f"  [skip] sheet '{sheet_name}' not found")
            continue

        ws = wb[sheet_name]
        mappings = {}
        csv_rows = []

        for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
            source_id = row[0]
            my_id = row[1]

            if source_id is None or my_id is None:
                continue

            source_id = str(source_id).strip()
            if isinstance(my_id, bool):
                my_id = str(my_id)
            else:
                try:
                    my_id = str(int(my_id))
                except (ValueError, TypeError):
                    my_id = str(my_id).strip()

            if not source_id or not my_id:
                continue

            if my_id in by_my_id:
                print(f"  [warn] duplicate my_id={my_id} (source={source_id})")

            mappings[source_id] = my_id
            csv_rows.append(f"{source_id},{my_id}")

            source_url_builder = SOURCE_URL_BUILDERS.get(oj_key)
            source_url = source_url_builder(source_id) if source_url_builder else None

            by_my_id[my_id] = {
                "source": oj_key,
                "source_id": source_id,
                "source_url": source_url,
            }

        by_source[oj_key] = mappings
        count = len(mappings)
        total += count

        csv_path = os.path.join(OUTPUT_CSV_DIR, f"{oj_key}.csv")
        with open(csv_path, "w", encoding="utf-8", newline="") as f:
            f.write("source_id,my_id\n")
            for line in csv_rows:
                f.write(line + "\n")

        print(f"  {oj_key} ({sheet_name}): {count} problems")

    wb.close()

    data = {
        "version": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "by_source": by_source,
        "by_my_id": by_my_id,
        "remote_prefixes": REMOTE_JUDGE_PREFIXES,
    }

    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"\nDone: {total} total mappings → {OUTPUT_JSON}, {OUTPUT_CSV_DIR}/")


if __name__ == "__main__":
    build()
